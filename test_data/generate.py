"""Generate test data files for dview — various formats, enough rows/cols for scroll testing."""

# /// script
# requires-python = ">=3.10"
# dependencies = [
#     "openpyxl",
#     "xlwt",
#     "pyarrow",
# ]
# ///

import csv
import random
import string
from pathlib import Path

OUT = Path("test_data")
OUT.mkdir(exist_ok=True)

ROWS = 100
COLS = 30

# ---------------------------------------------------------------------------
# Generate varied cell content
# ---------------------------------------------------------------------------
rng = random.Random(42)

FIRST_NAMES = [
    "Alice", "Bob", "Charlie", "Diana", "Eve", "Frank", "Grace",
    "Hank", "Iris", "Jack", "Kate", "Leo", "Mia", "Noah", "Olivia",
    "Paul", "Quinn", "Rose", "Sam", "Tina", "Uma", "Victor", "Wendy",
    "Xavier", "Yara", "Zack", "张三", "李四", "王五", "赵六",
]

CITIES = [
    "New York", "Los Angeles", "Chicago", "Houston", "Phoenix",
    "Beijing", "Shanghai", "Guangzhou", "Shenzhen", "Tokyo",
    "London", "Paris", "Berlin", "Sydney", "Mumbai",
]

DEPARTMENTS = [
    "Engineering", "Marketing", "Sales", "HR", "Finance",
    "Operations", "Legal", "Design", "Support", "R&D",
]

PRODUCTS = [
    "Widget A", "Widget B", "Gadget X", "Gadget Y", "Thingamajig Pro",
    "Doohickey Lite", "Contraption Max", "Whatchamacallit",
]

NOTES_POOL = [
    "",
    "Lorem ipsum dolor sit amet",
    "备注：需要进一步审核",
    "N/A",
    "Very long description that spans many characters to test horizontal scrolling behavior in dview",
    "特别长的中文描述文本用来测试水平滚动功能是否正常工作",
    "OK",
    "Pending approval from manager",
    "待审批",
    "Completed ✓",
]


def rand_name():
    return rng.choice(FIRST_NAMES)


def rand_city():
    return rng.choice(CITIES)


def rand_dept():
    return rng.choice(DEPARTMENTS)


def rand_product():
    return rng.choice(PRODUCTS)


def rand_note():
    return rng.choice(NOTES_POOL)


def rand_int(lo=0, hi=100000):
    return rng.randint(lo, hi)


def rand_float(lo=0.0, hi=10000.0):
    return round(rng.uniform(lo, hi), 2)


def rand_date():
    y = rng.randint(2020, 2026)
    m = rng.randint(1, 12)
    d = rng.randint(1, 28)
    return f"{y}-{m:02d}-{d:02d}"


def rand_email(name_col):
    # Use the name from the row; if CJK, use pinyin-ish fallback
    domain = rng.choice(["example.com", "company.org", "test.net", "邮件.cn"])
    user = name_col.lower().replace(" ", ".")
    return f"{user}@{domain}"


# Build varied column definitions
HEADERS = [
    "ID",
    "Name",
    "Age",
    "City",
    "Department",
    "Salary",
    "Bonus",
    "Start_Date",
    "End_Date",
    "Product",
    "Quantity",
    "Unit_Price",
    "Total",
    "Manager",
    "Email",
    "Phone",
    "Address",
    "Notes",
    "Score",
    "Rank",
    "Region",
    "Country",
    "Tax_Rate",
    "Net_Income",
    "Currency",
    "Active",
    "Projects",
    "Years_Exp",
    "Training_Hours",
    "Status",
]


def generate_row(i: int) -> list:
    """Generate one data row with mixed types and some wide CJK content."""
    name = rand_name()
    age = rand_int(18, 65)
    city = rand_city()
    dept = rand_dept()
    salary = rand_int(30000, 200000)
    bonus = rand_float(0, 50000)
    start_date = rand_date()
    end_date = rand_date()
    product = rand_product()
    qty = rand_int(1, 500)
    unit_price = rand_float(1.0, 999.99)
    total = round(qty * unit_price, 2)
    manager = rand_name()
    email = rand_email(name)
    phone = f"+86-{rng.randint(100, 999)}-{rng.randint(1000, 9999)}-{rng.randint(1000, 9999)}"
    address = f"{rand_int(1, 999)} {rng.choice(['Main St', 'Oak Ave', '长安街', '南京路', 'Park Rd'])}"
    notes = rand_note()
    score = rand_float(0, 100)
    rank = rng.choice(["A", "B", "C", "D", "S", "甲", "乙", "丙"])
    region = rng.choice(["North", "South", "East", "West", "东北", "华北", "华东", "西南"])
    country = rng.choice(["USA", "China", "Japan", "UK", "Germany", "France", "Australia", "中国", "日本"])
    tax_rate = round(rng.uniform(0.05, 0.45), 2)
    net_income = round(salary * (1 - tax_rate), 2)
    currency = rng.choice(["USD", "CNY", "EUR", "JPY", "GBP"])
    active = str(rng.choice([True, False])).lower()
    projects = rand_int(0, 20)
    years_exp = rand_int(0, 40)
    training_hours = rand_int(0, 200)
    status = rng.choice(["Active", "Inactive", "On Leave", "Terminated", "在职", "离职"])

    return [
        i + 1, name, age, city, dept, salary, bonus,
        start_date, end_date, product, qty, unit_price,
        total, manager, email, phone, address, notes,
        score, rank, region, country, tax_rate, net_income,
        currency, active, projects, years_exp, training_hours, status,
    ]


# ---------------------------------------------------------------------------
# Generate data
# ---------------------------------------------------------------------------
rows = [generate_row(i) for i in range(ROWS)]

# Convert all cells to string (dview loads everything as string anyway)
def as_str(row):
    return [str(c) for c in row]

str_rows = [as_str(r) for r in rows]


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------
with open(OUT / "sample.csv", "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(HEADERS)
    w.writerows(str_rows)
print(f"  sample.csv — {ROWS} rows × {COLS} cols")


# ---------------------------------------------------------------------------
# TSV
# ---------------------------------------------------------------------------
with open(OUT / "sample.tsv", "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f, delimiter="\t")
    w.writerow(HEADERS)
    w.writerows(str_rows)
print(f"  sample.tsv — {ROWS} rows × {COLS} cols")


# ---------------------------------------------------------------------------
# XLSX (single sheet)
# ---------------------------------------------------------------------------
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Data"
ws.append(HEADERS)
for r in str_rows:
    ws.append(r)
wb.save(OUT / "sample.xlsx")
print(f"  sample.xlsx (1 sheet) — {ROWS} rows × {COLS} cols")


# ---------------------------------------------------------------------------
# XLS (single sheet, old format)
# ---------------------------------------------------------------------------
import xlwt

wb_xls = xlwt.Workbook(encoding="utf-8")
ws_xls = wb_xls.add_sheet("Data")
for c, h in enumerate(HEADERS):
    ws_xls.write(0, c, h)
for ri, row in enumerate(str_rows):
    for ci, val in enumerate(row):
        ws_xls.write(ri + 1, ci, val)
wb_xls.save(str(OUT / "sample.xls"))
print(f"  sample.xls (1 sheet) — {ROWS} rows × {COLS} cols")


# ---------------------------------------------------------------------------
# XLSX with multiple sheets
# ---------------------------------------------------------------------------
wb_multi = Workbook()

# Sheet 1 — full data
ws1 = wb_multi.active
ws1.title = "All Data"
ws1.append(HEADERS)
for r in str_rows:
    ws1.append(r)

# Sheet 2 — summary
ws2 = wb_multi.create_sheet("Summary")
ws2.append(["Metric", "Value", "Unit", "Date", "Notes"])
summary_data = [
    ["Total Employees", str(ROWS), "people", "2026-07-01", "All departments"],
    ["Average Salary", f"{sum(int(r[5]) for r in rows)//ROWS}", "USD", "2026-07-01", "Annual"],
    ["Average Age", f"{sum(int(r[2]) for r in rows)//ROWS}", "years", "2026-07-01", ""],
    ["Total Revenue", str(sum(r[12] for r in rows)), "USD", "2026-07-01", "含税"],
    ["Active Rate", f"{sum(1 for r in rows if r[25] == 'true') / ROWS:.1%}", "", "2026-07-01", ""],
    ["中位数工资", "待计算", "CNY", "2026-07-01", "需要进一步核实"],
]
for row in summary_data:
    ws2.append(row)

# Sheet 3 — department breakdown
ws3 = wb_multi.create_sheet("By Department")
ws3.append(["Department", "Headcount", "Avg Salary", "Avg Score", "Budget"])
dept_data = {}
for r in rows:
    d = r[4]
    if d not in dept_data:
        dept_data[d] = {"count": 0, "salary": 0, "score": 0}
    dept_data[d]["count"] += 1
    dept_data[d]["salary"] += r[5]
    dept_data[d]["score"] += r[18]
for dept, data in sorted(dept_data.items()):
    ws3.append([
        dept,
        data["count"],
        round(data["salary"] / data["count"], 2),
        round(data["score"] / data["count"], 2),
        rand_int(500000, 5000000),
    ])

wb_multi.save(OUT / "sample_multi.xlsx")
print(f"  sample_multi.xlsx (3 sheets) — All Data / Summary / By Department")


# ---------------------------------------------------------------------------
# XLS with multiple sheets (old format)
# ---------------------------------------------------------------------------
wb_xls_multi = xlwt.Workbook(encoding="utf-8")

# Sheet 1
ws_x1 = wb_xls_multi.add_sheet("All Data")
for c, h in enumerate(HEADERS):
    ws_x1.write(0, c, h)
for ri, row in enumerate(str_rows):
    for ci, val in enumerate(row):
        ws_x1.write(ri + 1, ci, val)

# Sheet 2
ws_x2 = wb_xls_multi.add_sheet("Summary")
summary_headers = ["Metric", "Value", "Unit", "Date", "Notes"]
for c, h in enumerate(summary_headers):
    ws_x2.write(0, c, h)
for ri, row in enumerate(summary_data):
    for ci, val in enumerate(row):
        ws_x2.write(ri + 1, ci, val)

# Sheet 3
ws_x3 = wb_xls_multi.add_sheet("By Department")
dept_headers = ["Department", "Headcount", "Avg Salary", "Avg Score", "Budget"]
for c, h in enumerate(dept_headers):
    ws_x3.write(0, c, h)
for ri, (dept, data) in enumerate(sorted(dept_data.items())):
    ws_x3.write(ri + 1, 0, dept)
    ws_x3.write(ri + 1, 1, data["count"])
    ws_x3.write(ri + 1, 2, round(data["salary"] / data["count"], 2))
    ws_x3.write(ri + 1, 3, round(data["score"] / data["count"], 2))
    ws_x3.write(ri + 1, 4, rand_int(500000, 5000000))

wb_xls_multi.save(str(OUT / "sample_multi.xls"))
print(f"  sample_multi.xls (3 sheets) — All Data / Summary / By Department")


# ---------------------------------------------------------------------------
# Parquet
# ---------------------------------------------------------------------------
import pyarrow as pa
import pyarrow.parquet as pq

# Build pyarrow columns
col_arrays = {}
for ci, h in enumerate(HEADERS):
    col_vals = [str_rows[ri][ci] for ri in range(ROWS)]
    col_arrays[h] = pa.array(col_vals, type=pa.string())

table = pa.table(col_arrays)
pq.write_table(table, OUT / "sample.parquet")
print(f"  sample.parquet — {ROWS} rows × {COLS} cols")


# ---------------------------------------------------------------------------
# Done
# ---------------------------------------------------------------------------
print(f"\nAll files written to {OUT.resolve()}/")
